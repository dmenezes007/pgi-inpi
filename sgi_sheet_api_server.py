#!/usr/bin/env python3
"""
Serve static files and exposes an API backed by an XLSX workbook.

Run:
    python sgi_sheet_api_server.py

Optional env vars:
    INPI_IDEAS_XLSX -> absolute or relative path to workbook file
    INPI_SERVER_PORT -> server port (default: 8010)
    INPI_SERVER_HOST -> bind host (default: 0.0.0.0)
    INPI_CORS_ORIGIN -> allowed CORS origin (default: *)
    INPI_API_BASE_PATH -> base path for the API (default: /api)
"""

from __future__ import annotations

import json
import mimetypes
import os
from email.parser import BytesParser
from email.policy import default as email_default_policy
import threading
import uuid
from datetime import datetime
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import quote, urlparse

try:
    from openpyxl import Workbook, load_workbook
except Exception as exc:  # pragma: no cover
    raise RuntimeError(
        "openpyxl is required. Install with: pip install openpyxl"
    ) from exc


BASE_DIR = Path(__file__).resolve().parent
PORT = int(os.getenv("PORT", os.getenv("INPI_SERVER_PORT", "8010")))
HOST = os.getenv("INPI_SERVER_HOST", "0.0.0.0")
API_BASE_PATH = os.getenv("INPI_API_BASE_PATH", "/api").rstrip("/") or "/api"
ALLOWED_CORS_ORIGIN = os.getenv("INPI_CORS_ORIGIN", "*")
WORKBOOK_PATH = Path(os.getenv("INPI_IDEAS_XLSX", str(BASE_DIR / "inpi_ideas_database.xlsx")))
UPLOADS_DIR = Path(os.getenv("INPI_UPLOADS_DIR", str(BASE_DIR / "uploads")))
MAX_UPLOAD_SIZE_BYTES = int(os.getenv("INPI_MAX_UPLOAD_SIZE_BYTES", str(10 * 1024 * 1024)))

ALLOWED_UPLOAD_EXTENSIONS = {
    ".pdf",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".ppt",
    ".pptx",
    ".png",
    ".jpg",
    ".jpeg",
}

IDEAS_SHEET = "ideas"
LIFECYCLE_SHEET = "lifecycle"

IDEAS_HEADERS = [
    "id",
    "timestamp",
    "autor",
    "email",
    "titulo",
    "categoria",
    "problema",
    "justificativa",
    "escopo",
    "integrantes_json",
    "etapas_json",
    "recursos_json",
    "riscos_json",
    "indicadores_json",
    "status",
    "likes",
    "is_example",
    "anexos_json",
    "updated_at",
]

LIFECYCLE_HEADERS = [
    "idea_id",
    "status",
    "timestamp",
    "justificativa",
    "actor",
]

LIST_FIELDS = [
    "integrantes",
    "etapas",
    "recursos",
    "riscos",
    "indicadores",
    "anexos",
]

file_lock = threading.Lock()


def _now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def _ensure_workbook() -> None:
    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    if WORKBOOK_PATH.exists():
        wb = load_workbook(WORKBOOK_PATH)
    else:
        wb = Workbook()

    if IDEAS_SHEET not in wb.sheetnames:
        ws = wb.active
        ws.title = IDEAS_SHEET
        ws.append(IDEAS_HEADERS)
    else:
        ws = wb[IDEAS_SHEET]
        if ws.max_row == 0:
            ws.append(IDEAS_HEADERS)

    if LIFECYCLE_SHEET not in wb.sheetnames:
        lifecycle = wb.create_sheet(LIFECYCLE_SHEET)
        lifecycle.append(LIFECYCLE_HEADERS)
    else:
        lifecycle = wb[LIFECYCLE_SHEET]
        if lifecycle.max_row == 0:
            lifecycle.append(LIFECYCLE_HEADERS)

    wb.save(WORKBOOK_PATH)


def _api_path(suffix: str = "") -> str:
        suffix = suffix.strip()
        if suffix and not suffix.startswith("/"):
            suffix = "/" + suffix
        return f"{API_BASE_PATH}{suffix}"


def _parse_list(raw_value):
    if raw_value in (None, ""):
        return []
    try:
        parsed = json.loads(raw_value)
        return parsed if isinstance(parsed, list) else []
    except Exception:
        return []


def _to_bool(raw_value) -> bool:
    if isinstance(raw_value, bool):
        return raw_value
    if raw_value is None:
        return False
    return str(raw_value).strip().lower() in {"1", "true", "yes", "sim"}


def _sanitize_upload_name(filename: str) -> str:
    # Keep only leaf filename and strip dangerous whitespace.
    safe_name = Path(filename or "").name.strip()
    if not safe_name:
        raise ValueError("filename is required")
    return safe_name


def _store_upload(filename: str, file_bytes: bytes, content_type: str) -> dict:
    safe_name = _sanitize_upload_name(filename)
    extension = Path(safe_name).suffix.lower()

    if extension not in ALLOWED_UPLOAD_EXTENSIONS:
        raise ValueError("file extension not allowed")

    if not file_bytes:
        raise ValueError("empty file is not allowed")

    if len(file_bytes) > MAX_UPLOAD_SIZE_BYTES:
        raise ValueError("file exceeds maximum allowed size")

    file_id = uuid.uuid4().hex
    stored_name = f"{file_id}{extension}"
    target_path = UPLOADS_DIR / stored_name
    target_path.write_bytes(file_bytes)

    return {
        "id": file_id,
        "originalName": safe_name,
        "storedName": stored_name,
        "contentType": content_type or mimetypes.guess_type(safe_name)[0] or "application/octet-stream",
        "size": len(file_bytes),
        "uploadedAt": _now_iso(),
        "downloadUrl": f"/api/uploads/{stored_name}",
    }


def _resolve_upload_path(stored_name: str) -> Path:
    # Prevent path traversal and only allow files inside uploads directory.
    safe_leaf = Path(stored_name or "").name
    if not safe_leaf or safe_leaf != stored_name:
        raise ValueError("invalid upload id")

    path = (UPLOADS_DIR / safe_leaf).resolve()
    uploads_root = UPLOADS_DIR.resolve()
    if uploads_root not in path.parents and path != uploads_root:
        raise ValueError("invalid upload path")
    if not path.exists() or not path.is_file():
        raise FileNotFoundError("upload not found")
    return path


def _parse_multipart_payload(body: bytes, content_type: str) -> dict[str, list[dict]]:
    if "multipart/form-data" not in (content_type or "").lower():
        raise ValueError("multipart/form-data is required")

    message = BytesParser(policy=email_default_policy).parsebytes(
        f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8") + body
    )

    fields: dict[str, list[dict]] = {}
    if not message.is_multipart():
        return fields

    for part in message.iter_parts():
        content_disposition = part.get("Content-Disposition", "")
        if not content_disposition:
            continue

        field_name = part.get_param("name", header="Content-Disposition")
        if not field_name:
            continue

        fields.setdefault(field_name, []).append(
            {
                "filename": part.get_filename(),
                "content_type": part.get_content_type(),
                "content": part.get_payload(decode=True) or b"",
                "text": part.get_content() if part.get_content_maintype() == "text" else None,
            }
        )

    return fields


def _load_all_ideas() -> list[dict]:
    with file_lock:
        _ensure_workbook()
        wb = load_workbook(WORKBOOK_PATH)
        ws_ideas = wb[IDEAS_SHEET]
        ws_lifecycle = wb[LIFECYCLE_SHEET]

        lifecycle_by_id: dict[str, list[dict]] = {}
        for row in ws_lifecycle.iter_rows(min_row=2, values_only=True):
            idea_id, status, timestamp, justificativa, actor = row
            if not idea_id:
                continue
            lifecycle_by_id.setdefault(str(idea_id), []).append(
                {
                    "status": str(status or ""),
                    "timestamp": str(timestamp or _now_iso()),
                    "justificativa": str(justificativa or ""),
                    "actor": str(actor or ""),
                }
            )

        ideas: list[dict] = []
        for row in ws_ideas.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            idea = {
                "id": str(row[0]),
                "timestamp": str(row[1] or _now_iso()),
                "autor": str(row[2] or ""),
                "email": str(row[3] or ""),
                "titulo": str(row[4] or ""),
                "categoria": str(row[5] or ""),
                "problema": str(row[6] or ""),
                "justificativa": str(row[7] or ""),
                "escopo": str(row[8] or ""),
                "integrantes": _parse_list(row[9]),
                "etapas": _parse_list(row[10]),
                "recursos": _parse_list(row[11]),
                "riscos": _parse_list(row[12]),
                "indicadores": _parse_list(row[13]),
                "status": str(row[14] or "Enviada"),
                "likes": int(row[15] or 0),
                "isExample": _to_bool(row[16]),
                "anexos": _parse_list(row[17]),
                "lifecycle": lifecycle_by_id.get(str(row[0]), []),
            }

            if not idea["lifecycle"]:
                idea["lifecycle"] = [
                    {
                        "status": idea["status"],
                        "timestamp": idea["timestamp"],
                        "justificativa": "Registro inicial.",
                        "actor": "Sistema",
                    }
                ]

            ideas.append(idea)

        ideas.sort(key=lambda i: i.get("timestamp", ""), reverse=True)
        return ideas


def _append_idea(idea: dict) -> dict:
    with file_lock:
        _ensure_workbook()
        wb = load_workbook(WORKBOOK_PATH)
        ws_ideas = wb[IDEAS_SHEET]
        ws_lifecycle = wb[LIFECYCLE_SHEET]

        idea_id = str(idea.get("id", "")).strip()
        if not idea_id:
            raise ValueError("id is required")

        existing_ids = {str(row[0]) for row in ws_ideas.iter_rows(min_row=2, values_only=True) if row[0]}
        if idea_id in existing_ids:
            raise ValueError("id already exists")

        likes = int(idea.get("likes", 0) or 0)
        timestamp = str(idea.get("timestamp") or _now_iso())

        ws_ideas.append(
            [
                idea_id,
                timestamp,
                str(idea.get("autor", "")),
                str(idea.get("email", "")),
                str(idea.get("titulo", "")),
                str(idea.get("categoria", "")),
                str(idea.get("problema", "")),
                str(idea.get("justificativa", "")),
                str(idea.get("escopo", "")),
                json.dumps(idea.get("integrantes", []), ensure_ascii=False),
                json.dumps(idea.get("etapas", []), ensure_ascii=False),
                json.dumps(idea.get("recursos", []), ensure_ascii=False),
                json.dumps(idea.get("riscos", []), ensure_ascii=False),
                json.dumps(idea.get("indicadores", []), ensure_ascii=False),
                str(idea.get("status", "Enviada")),
                likes,
                bool(idea.get("isExample", False)),
                json.dumps(idea.get("anexos", []), ensure_ascii=False),
                _now_iso(),
            ]
        )

        lifecycle_items = idea.get("lifecycle", []) or []
        if not lifecycle_items:
            lifecycle_items = [
                {
                    "status": str(idea.get("status", "Enviada")),
                    "timestamp": timestamp,
                    "justificativa": "Registro inicial.",
                    "actor": "Sistema",
                }
            ]

        for lc in lifecycle_items:
            ws_lifecycle.append(
                [
                    idea_id,
                    str(lc.get("status", "")),
                    str(lc.get("timestamp", _now_iso())),
                    str(lc.get("justificativa", "")),
                    str(lc.get("actor", "")),
                ]
            )

        wb.save(WORKBOOK_PATH)
        return idea


def _like_idea(idea_id: str) -> dict:
    with file_lock:
        _ensure_workbook()
        wb = load_workbook(WORKBOOK_PATH)
        ws_ideas = wb[IDEAS_SHEET]

        for row_idx in range(2, ws_ideas.max_row + 1):
            if str(ws_ideas.cell(row=row_idx, column=1).value or "") == idea_id:
                current = int(ws_ideas.cell(row=row_idx, column=16).value or 0)
                ws_ideas.cell(row=row_idx, column=16, value=current + 1)
                ws_ideas.cell(row=row_idx, column=19, value=_now_iso())
                wb.save(WORKBOOK_PATH)
                return {"id": idea_id, "likes": current + 1}

    raise KeyError("idea not found")


def _update_status(idea_id: str, status: str, justificativa: str, actor: str) -> dict:
    with file_lock:
        _ensure_workbook()
        wb = load_workbook(WORKBOOK_PATH)
        ws_ideas = wb[IDEAS_SHEET]
        ws_lifecycle = wb[LIFECYCLE_SHEET]

        found = False
        for row_idx in range(2, ws_ideas.max_row + 1):
            if str(ws_ideas.cell(row=row_idx, column=1).value or "") == idea_id:
                ws_ideas.cell(row=row_idx, column=15, value=status)
                ws_ideas.cell(row=row_idx, column=19, value=_now_iso())
                found = True
                break

        if not found:
            raise KeyError("idea not found")

        ws_lifecycle.append([idea_id, status, _now_iso(), justificativa, actor])
        wb.save(WORKBOOK_PATH)

        return {
            "id": idea_id,
            "status": status,
            "justificativa": justificativa,
            "actor": actor,
        }


class SgiApiHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(BASE_DIR), **kwargs)

    def end_headers(self):
        self.send_header("Access-Control-Allow-Origin", ALLOWED_CORS_ORIGIN)
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PATCH, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Access-Control-Max-Age", "86400")
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate")
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(HTTPStatus.NO_CONTENT)
        self.end_headers()

    def do_GET(self):
        path = urlparse(self.path).path
        if path == _api_path("/health"):
            return self._handle_health()
        if path == _api_path("/ideas"):
            return self._handle_get_ideas()
        if path.startswith(_api_path("/uploads/")):
            stored_name = path.split("/", 3)[3]
            return self._handle_download_upload(stored_name)
        if path == "/":
            self.path = "/sgi-index.html"
        return super().do_GET()

    def do_POST(self):
        path = urlparse(self.path).path
        if path == _api_path("/ideas"):
            return self._handle_create_idea()
        if path == _api_path("/uploads"):
            return self._handle_upload_file()
        if path.startswith(_api_path("/ideas/")) and path.endswith("/like"):
            idea_id = path.split("/")[3]
            return self._handle_like_idea(idea_id)
        self._send_json(HTTPStatus.NOT_FOUND, {"error": "route not found"})

    def do_PATCH(self):
        path = urlparse(self.path).path
        if path.startswith(_api_path("/ideas/")) and path.endswith("/status"):
            idea_id = path.split("/")[3]
            return self._handle_update_status(idea_id)
        self._send_json(HTTPStatus.NOT_FOUND, {"error": "route not found"})

    def _read_json(self) -> dict:
        raw_length = self.headers.get("Content-Length", "0")
        try:
            length = int(raw_length)
        except ValueError:
            length = 0

        if length <= 0:
            return {}

        raw = self.rfile.read(length).decode("utf-8")
        return json.loads(raw) if raw else {}

    def _send_json(self, status: HTTPStatus, payload: dict):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(int(status))
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _handle_get_ideas(self):
        try:
            ideas = _load_all_ideas()
            self._send_json(HTTPStatus.OK, {"ideas": ideas, "count": len(ideas)})
        except Exception as exc:
            self._send_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"error": str(exc)})

    def _handle_health(self):
        try:
            self._send_json(
                HTTPStatus.OK,
                {
                    "status": "ok",
                    "service": "sgi-sheet-api",
                    "apiBasePath": API_BASE_PATH,
                    "workbookReady": WORKBOOK_PATH.exists(),
                    "uploadsDirReady": UPLOADS_DIR.exists(),
                    "timestamp": _now_iso(),
                },
            )
        except Exception as exc:
            self._send_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"error": str(exc)})

    def _handle_create_idea(self):
        try:
            payload = self._read_json()
            saved = _append_idea(payload)
            self._send_json(HTTPStatus.CREATED, {"idea": saved})
        except ValueError as exc:
            self._send_json(HTTPStatus.CONFLICT, {"error": str(exc)})
        except Exception as exc:
            self._send_json(HTTPStatus.BAD_REQUEST, {"error": str(exc)})

    def _handle_like_idea(self, idea_id: str):
        try:
            result = _like_idea(idea_id)
            self._send_json(HTTPStatus.OK, result)
        except KeyError as exc:
            self._send_json(HTTPStatus.NOT_FOUND, {"error": str(exc)})
        except Exception as exc:
            self._send_json(HTTPStatus.BAD_REQUEST, {"error": str(exc)})

    def _handle_update_status(self, idea_id: str):
        try:
            payload = self._read_json()
            status = str(payload.get("status", "")).strip()
            justificativa = str(payload.get("justificativa", "")).strip()
            actor = str(payload.get("actor", "Laboratório de Inovação")).strip()

            if not status:
                return self._send_json(HTTPStatus.BAD_REQUEST, {"error": "status is required"})
            if not justificativa:
                return self._send_json(HTTPStatus.BAD_REQUEST, {"error": "justificativa is required"})

            result = _update_status(idea_id, status, justificativa, actor)
            self._send_json(HTTPStatus.OK, result)
        except KeyError as exc:
            self._send_json(HTTPStatus.NOT_FOUND, {"error": str(exc)})
        except Exception as exc:
            self._send_json(HTTPStatus.BAD_REQUEST, {"error": str(exc)})

    def _handle_upload_file(self):
        try:
            raw_length = self.headers.get("Content-Length", "0")
            try:
                content_length = int(raw_length)
            except ValueError:
                content_length = 0

            # Multipart adds overhead; reserve +1MB headroom for boundaries/metadata.
            if content_length <= 0:
                return self._send_json(HTTPStatus.BAD_REQUEST, {"error": "empty request body"})
            if content_length > MAX_UPLOAD_SIZE_BYTES + 1024 * 1024:
                return self._send_json(HTTPStatus.REQUEST_ENTITY_TOO_LARGE, {"error": "request too large"})

            content_type = self.headers.get("Content-Type", "")
            body = self.rfile.read(content_length)
            form = _parse_multipart_payload(body, content_type)

            if "file" not in form or not form["file"]:
                return self._send_json(HTTPStatus.BAD_REQUEST, {"error": "field 'file' is required"})

            field = form["file"][0]

            filename = str(field.get("filename") or "")
            file_bytes = field.get("content", b"")
            uploaded = _store_upload(filename, file_bytes, field.get("content_type", ""))
            return self._send_json(HTTPStatus.CREATED, {"upload": uploaded})
        except ValueError as exc:
            self._send_json(HTTPStatus.BAD_REQUEST, {"error": str(exc)})
        except Exception as exc:
            self._send_json(HTTPStatus.BAD_REQUEST, {"error": str(exc)})

    def _handle_download_upload(self, stored_name: str):
        try:
            file_path = _resolve_upload_path(stored_name)
            ctype = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
            data = file_path.read_bytes()

            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", ctype)
            self.send_header("Content-Length", str(len(data)))
            self.send_header(
                "Content-Disposition",
                f"attachment; filename*=UTF-8''{quote(file_path.name)}",
            )
            self.end_headers()
            self.wfile.write(data)
        except FileNotFoundError:
            self._send_json(HTTPStatus.NOT_FOUND, {"error": "upload not found"})
        except ValueError as exc:
            self._send_json(HTTPStatus.BAD_REQUEST, {"error": str(exc)})
        except Exception as exc:
            self._send_json(HTTPStatus.BAD_REQUEST, {"error": str(exc)})


def run():
    _ensure_workbook()
    server = ThreadingHTTPServer((HOST, PORT), SgiApiHandler)
    print("=" * 64)
    print("SGI Sheet API server started")
    print(f"Host: {HOST}")
    print(f"URL: http://{HOST}:{PORT}/")
    print(f"App: http://{HOST}:{PORT}/sgi-index.html")
    print(f"Workbook: {WORKBOOK_PATH}")
    print("Endpoints:")
    print(f"  GET   {_api_path('/health')}")
    print(f"  GET   {_api_path('/ideas')}")
    print(f"  POST  {_api_path('/ideas')}")
    print(f"  POST  {_api_path('/uploads')}")
    print(f"  GET   {_api_path('/uploads/<storedName>')}")
    print(f"  POST  {_api_path('/ideas/<id>/like')}")
    print(f"  PATCH {_api_path('/ideas/<id>/status')}")
    print("=" * 64)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    run()
